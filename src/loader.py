import pandas as pd
import math
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
import os

def loadModel(file_name, type='excel'):
    """
    Charge un model depuis un fichier excel ou csv.
    Le fichier excel doit respecter le format fournit en exemple

    Args:
        file_name: le nom du fichier placé dans le répertoire data
        type: le type de fichier ('excel' ou 'csv')
    Return:
        le fichier découpé selon les différents éléments
    """
    if type == 'csv':
        df = pd.read_csv(f'{file_name}')
    elif type == 'excel':
        df = pd.read_excel(f'{file_name}')
        
    return df

def parseDataframe(df):
    """
    Découpe le fichier en élements utiles

    Args:
        df: un DataFrame

    Return:
        df_score : la colonne score
        coeff_list : la liste des coefficients
        df_criteria_list : un Dataframe avec les colonnes produits 
            et les colonnes de critères
        df_criteria_bareme : un Dataframe avec les colonnes de critères
            et leurs bornes respectives por chaque lignes
        dict_boundaries : un Dict avec les valeurs min et max du barème
        df_criteria_profils : liste des profils avec les bornes correspondantes
        score_bymagazine : liste des scores sous forme de classes
        dict_categories: dict des labels des catégories
        
    """
    
    df_score  = df['Score'].copy()
    coeff_list = buildCoeffList(df)
    df_bareme = df[['Note','Min_value','Max_value']].copy()
    
    df_criteria_list = buildCriteriaList(df)
    
    df_criteria_profils=buildProfils(df,df_criteria_list)
    df_criteria_bareme = buildCriteriaBaremedf(df_criteria_list, df_bareme)
    
    dict_boundaries = {}
    dict_boundaries['min'] = df_bareme['Min_value'].min()
    dict_boundaries['max'] = df_bareme['Max_value'].max()
    
    score_bymagazine= buildScoreByMagazine(df)
    dict_categories = buildDictCategorie(df)

    return df_score, coeff_list, df_criteria_list, df_criteria_bareme, dict_boundaries, df_criteria_profils, score_bymagazine, dict_categories

def buildDictCategorie(df):
    """
    Renvoie un dataframe concaténant les colonnes Categorie et Valeurs
        
    Args:
        df : Dataframe original
    
    Return: Dataframe
    """
    cat_df = df[['Categories','Valeurs']].copy().dropna()
    cat_dict = dict(zip(cat_df['Categories'], cat_df['Valeurs']))
    return cat_dict
     
def buildCriteriaList(df):
    """
    Renvoie un dataframe contenant les critères et la colonne produits
        
    Args:
        df : Dataframe original
    
    Return: Dataframe
    """
    df_criteria_list = df.copy()
    for val in df_criteria_list.columns:
        if val.endswith('_Profil'):
            df_criteria_list = df_criteria_list.drop(val, axis = 1)
    return df_criteria_list.drop(['Score', 'Coefficient','Note','Min_value','Max_value','Profil', 'Bornes', 'Note_magazine', 'Categories', 'Valeurs'], 1)
                
                
def buildCriteriaBaremedf(df_criteria_list, df_bareme):
    """
    Pour chaque colonne de critères, rajoute les bornes min et max de 
        la valeur qualitative du critère
        
    Args:
        df_criteria_list: un DataFrame avec les valeurs quantitatives des critères
        df_bareme: un Dataframe avec les bornes min et max (le barème)
    
    Return: un Dataframe avec les colonnes de critères
            et leurs bornes respectives por chaque lignes
    """
    
    df_criteria_bareme = df_criteria_list.copy()
    
    for criteria in df_criteria_list.iloc[:,1:]:
        df_criteria_bareme = pd.merge(df_criteria_bareme, df_bareme, how='left', left_on=criteria, right_on=['Note'])
        df_criteria_bareme = df_criteria_bareme.drop(['Note'],1)
        df_criteria_bareme = df_criteria_bareme.rename(columns={'Min_value': f'Min_value_{criteria}', 'Max_value': f'Max_value_{criteria}'})
    
    return df_criteria_bareme

def buildCoeffList(df):
    """
    Récupère la liste des coefficients
        
    Args:
        df: Le Dataframe correspondant au fichier d'entrée
    
    Return: une liste de coefficient
    """
     
    coeff = df['Coefficient']
    coeff_list = []
    for c in coeff:
        if not math.isnan(c):
            coeff_list.append(c)
            
    return coeff_list

def buildProfils(df, df_criteria_list):
    
    columns_list = ['Profil']+[col+'_Profil' for col in df_criteria_list.columns.copy() if col != 'Produit']
    df_Profil_list = pd.DataFrame(columns=columns_list)
    for col in df_Profil_list:
        if col == 'Profil':
            df_Profil_list[col] = df['Profil'].copy()
        else:
            df_Profil_list[col] = df['Bornes'].copy()
    return df_Profil_list.dropna()
   

def buildScoreByMagazine(df):
    """
    Récupère la liste des scores obtenues par le magazine
        
    Args:
        df: Le Dataframe correspondant au fichier d'entrée
    
    Return: une liste de scores (catégories)
    """
    score = df['Note_magazine']
    score_bymagazine = []
    for s in score:
        if s!="Nan":
            score_bymagazine.append(s)
            
    return score_bymagazine
    

def getOriginalData(csv_name, type='excel'):
    """
    renvoie le Dataframe original du fichier
        
    Args:
        df: Le path du fichier
    Return: Dataframe
    """
    df = loadModel(csv_name, type)
    df_score  = df['Score'].copy()
    df_criteria_list = buildCriteriaList(df)
    df = df_criteria_list.join(df_score)
    return df

def getElectreTriData(csv_name, type='excel'):
    """
    renvoie les colonnes Produits et Note_magazine
        
    Args:
        df: Le path du fichier
    Return: Dataframe
    """
    df = loadModel(csv_name, type)
    df_produit  = df[['Produit', 'Note_magazine']]
    df = df_produit
    return df

def exportInExcel(filename, sheetname, df_list, name_list, rankings_list, electre=False):
    """
    Exporte les résulats sous forme de fichier excel
        
    Args:
        filename: Le path du fichier
        sheetname: Le nom de l'onglet excel
        df_list: liste des dataframes à exporter
        name_list: Liste des noms de chaque dataframe
        rankings_list: Liste des résultats à exporter
        electre: export Electre tri par défaut à False
        
    Return: un objet Workbook
    """
    #si le fichier n'existe pas on le cree
    if not os.path.isfile(filename):
        workbook = Workbook()
        workbook.save(filename)
        
    workbook = load_workbook(filename, read_only=False)
    #Si l'onglet existe on réécrit par dessus
    if 'Sheet' in workbook.sheetnames:
        delete_sheet = workbook.get_sheet_by_name('Sheet')
        workbook.remove_sheet(delete_sheet) 
    #Si l'onglet n'existe pas on le cree
    if sheetname not in workbook.sheetnames:
         workbook.create_sheet(sheetname)
         
    workbook = writeInWorkbook(workbook, sheetname, df_list, name_list, rankings_list, electre)        
    workbook.save(filename)
    return workbook

def writeInWorkbook(workbook, sheetname, df_list, name_list, rankings_list, electre=False):
    """
    Ecrit les résultats dans un fichier excel
        
    Args:
        workbook: le fichier excel dans un objet Workbook
        sheetname: Le nom de l'onglet excel
        df_list: liste des dataframes à exporter
        name_list: Liste des noms de chaque dataframe
        rankings_list: Liste des résultats à exporter
        electre: export Electre tri par défaut à False
        
    Return: un objet Workbook
    """
    
    sheet = workbook[sheetname]
    thin_border = thinBorders()   
     
    start_line = 2
    start_col = 2
    for idx, df in enumerate(df_list):
        
        #Pourchaque dataframe, on écrit le contenu dans le fichier
        rows = dataframe_to_rows(df, index=False, header=True)           
        start_cell(sheet, start_line, start_col, name_list, idx, thin_border)
        for r_idx, row in enumerate(rows, start_line+1):
            for c_idx, value in enumerate(row, start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value).border = thin_border 
        
        min_row = start_line+1
        max_row = r_idx
        min_col = start_col
        max_col = c_idx
        
        #Résultats et indicateurs
        if not electre:
            if idx !=0:
                indicators(rankings_list, idx, sheet, max_row, min_col, thin_border)
        else:
            print(rankings_list)
            indicators(rankings_list, idx, sheet, max_row, min_col, thin_border)
        
        #Style
        blue_col(sheet, min_row, max_row, min_col, thin_border)
        black_row(sheet, min_col, max_col, min_row, thin_border)
        columns_width(sheet)
            
        start_col = max_col+3
            
    return workbook

def thinBorders():
    """
    Crée les bordures des cases
    Return: un objet Border
    """
    thin_border = Border(left=Side(style='thin'), 
             right=Side(style='thin'), 
             top=Side(style='thin'), 
             bottom=Side(style='thin'))
    return thin_border

def indicators(rankings_list, idx, sheet, max_row, min_col, thin_border):
    """
    Met en page les différents indicateurs dans l'onglet excel
        
    Args:
        rankings_list: Liste des résultats à exporter
        idx: numero du Dataframe
        sheet: onglet Excel
        max_row: dernière ligne du Dataframe
        min_col: première colonne du Dataframe
        thin_border: objet border
   
    """
    
    rankings = rankings_list[idx-1]
    i = 3
    for key, value in rankings.items():
        cell_key = sheet.cell(row=max_row+i, column=min_col, value=key)
        cell_key.border = thin_border 
        cell_key.fill = PatternFill("solid", fgColor="04024F")
        cell_key.font = Font(bold=True, color="FFFFFF")
        cell_val = sheet.cell(row=max_row+i, column=min_col+1, value=value).border = thin_border
        cell_val.border = thin_border 
        i += 1

def start_cell(sheet, start_line, start_col, name_list, idx, thin_border):
    """
    Calcul la cellule de départ pour chaque dataframe
        
    Args:
        sheet: onglet Excel
        start_line: ligne de départ
        start_col: colonne de départ
        name_list: liste des titres à affecter à chaque tableu
        idx: indice du dataframe
        thin_border: objet border
   
    """
    
    start_cell = sheet[get_column_letter(start_col)+str(start_line)]
    sheet.merge_cells(get_column_letter(start_col)+str(start_line)+':'+
                              get_column_letter(start_col+3)+str(start_line))
    start_cell.font = Font(bold=True)
    sheet.cell(start_line, start_col, value=name_list[idx])


def blue_col(sheet, min_row, max_row, min_col, thin_border):
    """
    Met en page les colonnes des tableaux
        
    Args:
        sheet: onglet Excel
        min_row: ligne de départ
        max_row: dernière ligne du tableau
        min_col: colonne à coloriser
        thin_border: objet border
   
    """
    for row in range(min_row+1,max_row+1):
        cell = sheet[get_column_letter(min_col)+str(row)]
        cell.style = 'Pandas'
        cell.fill = PatternFill("solid", fgColor="04024F")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
def black_row(sheet, min_col, max_col, min_row, thin_border):
    """
    Met en page les headers de tableau
        
    Args:
        sheet: onglet Excel
        min_col: colonne de départ
        max_col: dernière colonne du tableau
        min_row: ligne de départ
        thin_border: objet border
   
    """
    for col in range(min_col,max_col+1):
        cell = sheet[get_column_letter(col)+str(min_row)]
        cell.style = 'Pandas'
        cell.fill = PatternFill("solid", fgColor="000000")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border

def columns_width(sheet):
    """
    Ajuste la largeur des colonnes
    """
    for i, col in enumerate(sheet.columns):
        sheet.column_dimensions[get_column_letter(i+1)].width = 20
    
    
    